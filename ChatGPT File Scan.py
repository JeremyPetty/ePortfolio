import openai
import csv
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

# Set your OpenAI API key here
openai.api_key = 'Enter Key'

def read_excel_data(excel_path):
    wb = load_workbook(filename=excel_path, read_only=True)
    sheet = wb.active
    data = []
    header = None

    for row in sheet.iter_rows(values_only=True):
        # Assuming the header is in the first row
        if header is None:
            header = row
            continue

        # Assuming the columns are in the same order as in the header
        item, product_name, bag_size, price_bag, pallet_price = row[:5]

        # Format the data as a tab-separated string
        row_data = f"{item}\t{product_name}\t{bag_size}\t{price_bag}\t{pallet_price}\t"
        data.append(row_data)

    return '\n'.join(data)

def analyze_text_with_chatgpt(text):
    response = openai.Completion.create(
        engine="text-davinci-002",
        prompt=text,
        max_tokens=1575 # Adjust this as needed
    )
    return response.choices[0].text

def process_excel():
    excel_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if excel_path:
        excel_data = read_excel_data(excel_path)
        analyzed_data = analyze_text_with_chatgpt(excel_data)
        result_text.delete(1.0, tk.END)
        result_text.insert(tk.END, analyzed_data)

def save_to_csv():
    data_to_save = result_text.get(1.0, tk.END)
    csv_filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Files", "*.csv")])
    if csv_filename:
        write_to_csv(data_to_save, csv_filename)

def write_to_csv(data, csv_filename):
    with open(csv_filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Document Information'])  # Header
        writer.writerow([data])  # Data

# Create the main window
root = tk.Tk()
root.title("Excel Text Analysis")

# Create buttons
process_button = tk.Button(root, text="Process Excel", command=process_excel)
save_button = tk.Button(root, text="Save to CSV", command=save_to_csv)

# Create result text box
result_text = tk.Text(root, height=10, width=80)

# Pack widgets
process_button.pack()
save_button.pack()
result_text.pack()

# Start the GUI main loop
root.mainloop()